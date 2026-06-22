"""
PDF → Excel Converter  |  Standalone desktop tool
==================================================
Double-click to run (no terminal needed).
Requires Python 3.8+.  All other dependencies are installed automatically.

To build a true single-file executable (no Python required on target machine):
    pip install pyinstaller
    pyinstaller --onefile --windowed pdf_to_excel_converter.py
This produces:
    Windows : dist/pdf_to_excel_converter.exe
    macOS   : dist/pdf_to_excel_converter   (or wrap in .app via --windowed)
    Linux   : dist/pdf_to_excel_converter
"""

# ── 0. Dependency bootstrap ────────────────────────────────────────────────────
import sys
import subprocess
import importlib

_REQUIRED = [
    ("pdfplumber", "pdfplumber"),
    ("openpyxl",   "openpyxl"),
]

def _pip_install(pkg: str):
    """Install a package, handling PEP 668 (externally-managed Python on Homebrew/Debian)."""
    base_cmd = [sys.executable, "-m", "pip", "install", pkg, "--quiet"]
    try:
        subprocess.check_call(base_cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except subprocess.CalledProcessError:
        # PEP 668: system Python refuses pip without --break-system-packages.
        # Prefer a user install first; fall back only if that also fails.
        try:
            subprocess.check_call(
                base_cmd + ["--user"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )
        except subprocess.CalledProcessError:
            subprocess.check_call(
                base_cmd + ["--break-system-packages"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
            )


def _ensure_deps():
    missing = []
    for pkg, imp in _REQUIRED:
        try:
            importlib.import_module(imp)
        except ImportError:
            missing.append(pkg)

    if not missing:
        return

    # Show a minimal Tk splash while installing
    try:
        import tkinter as tk
        splash = tk.Tk()
        splash.title("PDF → Excel")
        splash.resizable(False, False)
        tk.Label(
            splash,
            text=f"Installing required libraries:\n{', '.join(missing)}\n\nThis happens only once…",
            padx=30, pady=20, font=("Helvetica", 12),
        ).pack()
        splash.update()
    except Exception:
        splash = None

    for pkg in missing:
        _pip_install(pkg)

    if splash:
        splash.destroy()

_ensure_deps()

# ── 1. Imports (all deps now guaranteed) ──────────────────────────────────────
import os
import threading
import platform
import subprocess as _sp
import traceback
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── 2. Extraction logic ───────────────────────────────────────────────────────

def _clean_row(row):
    """Replace None cells with empty string, normalise whitespace."""
    return [
        " ".join(str(c).split()) if c is not None else ""
        for c in row
    ]


def _score_tables(tables):
    """Total data rows across all tables (more = better extraction)."""
    return sum(max(0, len(t) - 1) for t in tables if t)


def _detect_columns(header_words):
    """
    Given the header words (sorted by x0), cluster consecutive words that are
    very close together (part of the same multi-word column label, e.g. "Chq No/Ref No"),
    then return:
      - col_anchors  : list of x0 positions (one per logical column)
      - col_boundaries: list of x midpoints BETWEEN clusters (len = anchors - 1)
    Boundary is placed midway between the right edge of one cluster and the
    left edge of the next — so words that belong to a multi-word header label
    stay in their own column.
    """
    if not header_words:
        return [], []

    sorted_words = sorted(header_words, key=lambda w: w["x0"])

    # Compute all inter-word gaps (x0_next - x1_prev)
    gaps = []
    for i in range(len(sorted_words) - 1):
        gap = sorted_words[i + 1]["x0"] - sorted_words[i]["x1"]
        gaps.append(gap)

    if not gaps:
        return [sorted_words[0]["x0"]], []

    # Threshold: words within 40% of the median gap belong to the same label
    gaps_sorted = sorted(gaps)
    median_gap = gaps_sorted[len(gaps_sorted) // 2]
    split_threshold = max(median_gap * 0.4, 4)

    # Build clusters
    clusters = [[sorted_words[0]]]
    for i in range(len(sorted_words) - 1):
        if gaps[i] < split_threshold:
            clusters[-1].append(sorted_words[i + 1])
        else:
            clusters.append([sorted_words[i + 1]])

    # Column anchors = leftmost x0 of each cluster
    col_anchors = [min(w["x0"] for w in c) for c in clusters]

    # Column boundaries = midpoint between right edge of cluster[i] and left edge of cluster[i+1]
    col_boundaries = []
    for i in range(len(clusters) - 1):
        right_edge = max(w["x1"] for w in clusters[i])
        left_edge  = min(w["x0"] for w in clusters[i + 1])
        col_boundaries.append((right_edge + left_edge) / 2)

    return col_anchors, col_boundaries


def _extract_by_words(page):
    """
    Reconstruct a table from word bounding boxes.
    1. Group words into horizontal bands (rows) by vertical proximity.
    2. Find the header band using keyword matching, then cluster its words
       into logical columns (merges multi-word header labels like "Chq No/Ref No").
    3. Derive column boundaries from midpoints between column anchors.
    4. Assign every word in every band to its nearest column.
    5. Merge continuation bands (date column empty) into the previous row.
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=2,
                                keep_blank_chars=False, use_text_flow=False)
    if not words:
        return []

    # ── Step 1: band by vertical proximity ───────────────────────────────────
    bands = []
    for w in sorted(words, key=lambda x: (x["top"], x["x0"])):
        placed = False
        for band in bands:
            if abs(w["top"] - band["top"]) < 5:
                band["words"].append(w)
                placed = True
                break
        if not placed:
            bands.append({"top": w["top"], "words": [w]})

    if not bands:
        return []

    # ── Step 2: find header band ──────────────────────────────────────────────
    _HEADER_KEYWORDS = {"date", "particulars", "description", "narration",
                        "withdrawal", "debit", "deposit", "credit",
                        "balance", "amount", "ref", "chq", "txn", "details"}
    header_band = None
    for band in sorted(bands, key=lambda b: b["top"]):
        band_text = {w["text"].lower() for w in band["words"]}
        if len(band_text & _HEADER_KEYWORDS) >= 2:
            header_band = band
            break

    if header_band is None:
        for band in sorted(bands, key=lambda b: b["top"]):
            if len(band["words"]) >= 3:
                spread = (max(w["x0"] for w in band["words"]) -
                          min(w["x0"] for w in band["words"]))
                if spread > page.width * 0.3:
                    header_band = band
                    break

    if header_band is None:
        header_band = bands[0]

    col_anchors, boundaries = _detect_columns(header_band["words"])
    if not col_anchors:
        return []

    # ── Step 3: column count / assignment ─────────────────────────────────────
    num_cols = len(col_anchors)

    def assign_col(x0):
        for i, b in enumerate(boundaries):
            if x0 < b:
                return i
        return num_cols - 1

    # ── Step 4: build rows ─────────────────────────────────────────────────────
    result_rows = []
    for band in sorted(bands, key=lambda b: b["top"]):
        cells = [""] * num_cols
        for w in sorted(band["words"], key=lambda x: x["x0"]):
            ci = assign_col(w["x0"])
            cells[ci] += (" " if cells[ci] else "") + w["text"]
        result_rows.append([c.strip() for c in cells])

    # ── Step 5: merge continuation rows ───────────────────────────────────────
    half = max(1, num_cols // 2)
    merged = []
    for row in result_rows:
        non_empty_cols = [i for i, v in enumerate(row) if v.strip()]
        is_continuation = (
            merged and
            not row[0].strip() and
            all(i < half + 1 for i in non_empty_cols)
        )
        if is_continuation:
            for ci in non_empty_cols:
                if ci < len(merged[-1]):
                    merged[-1][ci] = (merged[-1][ci] + " " + row[ci]).strip()
        else:
            merged.append(list(row))

    return merged


def extract_tables_from_pdf(pdf_path: str, progress_cb=None, log_cb=None):
    """
    Returns a list of dicts:
        { "label": str, "rows": [[str, ...], ...] }

    Strategy per page (best result wins):
      1. Line-snapping  — pdfplumber default  (solid-border tables)
      2. Word-grouping  — column anchors from header row  (borderless tables)
    Fallback: raw text lines when no tables are found.
    """
    results = []
    log = log_cb or (lambda msg: None)

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)
        log(f"Opened PDF  —  {total} page(s)")

        for page_idx, page in enumerate(pdf.pages, start=1):
            if progress_cb:
                progress_cb(page_idx, total)

            # ── Pass 1: line-detected / bordered tables ───────────────────────
            bordered = []
            try:
                for tbl in page.extract_tables():
                    if tbl:
                        rows = [_clean_row(r) for r in tbl if any(c for c in r if c)]
                        if rows:
                            bordered.append(rows)
            except Exception:
                pass

            # ── Pass 2: word-grouped ──────────────────────────────────────────
            word_rows = _extract_by_words(page)
            word_based = [word_rows] if word_rows else []

            # Keep whichever strategy has more data rows
            if _score_tables(word_based) > _score_tables(bordered):
                best = word_based
                strategy = "words"
            elif bordered:
                best = bordered
                strategy = "lines"
            else:
                best = word_based
                strategy = "words"

            best = [t for t in best if t]

            # ── Fallback: raw text ────────────────────────────────────────────
            if not best:
                raw = page.extract_text()
                if raw and raw.strip():
                    lines = [line for line in raw.splitlines() if line.strip()]
                    if lines:
                        best = [[[line] for line in lines]]
                        log(f"  Page {page_idx}: no tables — saved {len(lines)} raw-text lines")

            for tbl_idx, rows in enumerate(best, start=1):
                label = (
                    f"Page{page_idx}_Table{tbl_idx}"
                    if len(best) > 1
                    else f"Page{page_idx}"
                )
                results.append({"label": label, "rows": rows})
                log(
                    f"  Page {page_idx}: table {tbl_idx}  —  "
                    f"{len(rows)} rows × {len(rows[0])} cols  [{strategy}]"
                )

            if not best:
                log(f"  Page {page_idx}: (empty / image-only, skipped)")

    if len(results) == 1:
        results[0]["label"] = "Table"

    return results


def _looks_like_data_header(row):
    """True if the row looks like a real column-header row (not a cover-page title)."""
    if not row:
        return False
    non_empty = [c for c in row if c.strip()]
    if len(non_empty) < 2:
        return False
    _KW = {"date", "particulars", "description", "narration",
           "withdrawal", "debit", "deposit", "credit",
           "balance", "amount", "ref", "chq", "txn", "details"}
    first_tokens = {c.lower().strip().split()[0] for c in non_empty if c.strip()}
    return bool(first_tokens & _KW)


def _norm_cell(cell):
    """First token, lowercased — handles bleed like 'Particulars otil' → 'particulars'."""
    s = cell.lower().strip()
    return s.split()[0] if s else ""


def _all_headers_match(tables):
    """Return True when ≥2 real data tables all share the same (normalised) header."""
    real = [t for t in tables if t.get("rows") and _looks_like_data_header(t["rows"][0])]
    if len(real) < 2:
        return False
    first = [_norm_cell(c) for c in real[0]["rows"][0]]
    return all([_norm_cell(c) for c in t["rows"][0]] == first for t in real[1:])


def _combine_tables(tables):
    """Merge data tables into one sheet; keep non-data sheets (e.g. cover page) separate."""
    real = [t for t in tables if t.get("rows") and _looks_like_data_header(t["rows"][0])]
    other = [t for t in tables if not (t.get("rows") and _looks_like_data_header(t["rows"][0]))]
    header = real[0]["rows"][0]
    data = [row for t in real for row in t["rows"][1:]]
    return other + [{"label": "Combined", "rows": [header] + data}]


# ── 3. Excel writing ──────────────────────────────────────────────────────────

_HDR_FILL  = PatternFill("solid", fgColor="2E4057")
_HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
_ODD_FILL  = PatternFill("solid", fgColor="F5F7FA")
_EVEN_FILL = PatternFill("solid", fgColor="FFFFFF")
_THIN      = Side(style="thin", color="D0D0D0")
_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_SUMM_FILL = PatternFill("solid", fgColor="1A73E8")
_SUMM_FONT = Font(bold=True, color="FFFFFF", size=12)


def _auto_col_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                # don't let a single cell blow out the column
                max_len = max(max_len, min(len(val), 60))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)


def write_excel(tables, pdf_path: str, output_path: str):
    wb = openpyxl.Workbook()

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"] = "PDF → Excel Conversion Summary"
    ws_sum["A1"].fill = _SUMM_FILL
    ws_sum["A1"].font = _SUMM_FONT
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    summary_rows = [
        ("Source file",   Path(pdf_path).name),
        ("Converted on",  datetime.now().strftime("%d %b %Y  %H:%M")),
        ("Tables found",  len(tables)),
        ("Output sheets", ", ".join(t["label"] for t in tables) or "—"),
    ]
    for i, (key, val) in enumerate(summary_rows, start=2):
        ws_sum.cell(i, 1, key).font  = Font(bold=True)
        ws_sum.cell(i, 2, str(val))
        for col in (1, 2):
            ws_sum.cell(i, col).border = _BORDER
            ws_sum.cell(i, col).alignment = Alignment(vertical="center", wrap_text=True)

    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 60

    # ── One sheet per table ──────────────────────────────────────────────────
    for tbl in tables:
        label = tbl["label"][:31]   # Excel sheet name limit
        rows  = tbl["rows"]
        ws    = wb.create_sheet(title=label)

        for r_idx, row in enumerate(rows, start=1):
            is_header = (r_idx == 1)
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(r_idx, c_idx, val)
                if is_header:
                    cell.fill = _HDR_FILL
                    cell.font = _HDR_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                else:
                    cell.fill = _ODD_FILL if r_idx % 2 else _EVEN_FILL
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = _BORDER

        ws.freeze_panes = "A2"
        _auto_col_width(ws)

    wb.save(output_path)


# ── 4. Output path helper ─────────────────────────────────────────────────────

def _output_path(pdf_path: str) -> str:
    p    = Path(pdf_path)
    base = p.parent / (p.stem + "_tables")
    out  = base.with_suffix(".xlsx")
    n    = 1
    while out.exists():
        out = base.parent / f"{base.name}_{n}.xlsx"
        n  += 1
    return str(out)


def _open_file(path: str):
    """Open a file with the OS default application."""
    system = platform.system()
    if system == "Darwin":
        _sp.Popen(["open", path])
    elif system == "Windows":
        os.startfile(path)           # noqa: F405
    else:
        _sp.Popen(["xdg-open", path])


# ── 5. GUI ────────────────────────────────────────────────────────────────────

# TkinterDnD.Tk must be the root window class for drag-and-drop to function;
# registering DnD on a plain tk.Tk window silently does nothing or breaks events.
try:
    from tkinterdnd2 import TkinterDnD as _TkinterDnD, DND_FILES as _DND_FILES
    _AppBase = _TkinterDnD.Tk
    _DND_AVAILABLE = True
except Exception:
    _AppBase = tk.Tk
    _DND_FILES = None
    _DND_AVAILABLE = False


class App(_AppBase):
    _BG      = "#F0F4F8"
    _ACCENT  = "#1A73E8"
    _BTN_FG  = "#FFFFFF"
    _RADIUS  = 6

    def __init__(self):
        super().__init__()
        self.title("PDF → Excel Converter")
        self.configure(bg=self._BG)
        self.resizable(True, True)
        self.minsize(560, 480)

        self._pdf_path   = tk.StringVar()
        self._status_msg = tk.StringVar(value="Select a PDF file to get started.")
        self._output_path = None
        self._converting  = False

        self._build_ui()
        self._center_window(600, 520)

        if _DND_AVAILABLE:
            self.drop_target_register(_DND_FILES)
            self.dnd_bind("<<Drop>>", self._on_drop)
            self._log("Drag-and-drop enabled.")

    # ── Layout ────────────────────────────────────────────────────────────────

    def _build_ui(self):
        pad = dict(padx=16, pady=6)

        # Title bar
        title_frm = tk.Frame(self, bg=self._ACCENT)
        title_frm.pack(fill="x")
        tk.Label(
            title_frm, text="  PDF → Excel Converter",
            bg=self._ACCENT, fg="white",
            font=("Helvetica", 14, "bold"), anchor="w", pady=10,
        ).pack(fill="x", padx=12)

        body = tk.Frame(self, bg=self._BG)
        body.pack(fill="both", expand=True, padx=20, pady=12)

        # Drop / Browse zone
        drop_frm = tk.Frame(
            body, bg="#DDEEFF", relief="groove", bd=2,
            cursor="hand2",
        )
        drop_frm.pack(fill="x", pady=(0, 10))
        drop_frm.bind("<Button-1>", lambda _: self._browse())

        drop_lbl = tk.Label(
            drop_frm,
            text="📂  Drop a PDF here  or  click to Browse",
            bg="#DDEEFF", fg="#2C5282",
            font=("Helvetica", 12), pady=18,
            cursor="hand2",
        )
        drop_lbl.pack()
        drop_lbl.bind("<Button-1>", lambda _: self._browse())

        # Selected file label
        tk.Label(body, text="Selected file:", bg=self._BG,
                 font=("Helvetica", 10, "bold"), anchor="w").pack(fill="x")
        self._file_label = tk.Label(
            body, textvariable=self._pdf_path,
            bg=self._BG, fg="#444", font=("Helvetica", 10),
            anchor="w", wraplength=520,
        )
        self._file_label.pack(fill="x", pady=(0, 8))

        # Progress bar
        self._progress = ttk.Progressbar(body, mode="determinate", maximum=100)
        self._progress.pack(fill="x", pady=(0, 4))

        # Status
        tk.Label(
            body, textvariable=self._status_msg,
            bg=self._BG, fg="#555", font=("Helvetica", 10), anchor="w",
        ).pack(fill="x", pady=(0, 8))

        # Buttons
        btn_frm = tk.Frame(body, bg=self._BG)
        btn_frm.pack(fill="x", pady=(0, 8))

        self._btn_convert = tk.Button(
            btn_frm, text="  Convert  ",
            bg=self._ACCENT, fg=self._BTN_FG,
            font=("Helvetica", 11, "bold"),
            relief="flat", cursor="hand2", padx=14, pady=6,
            command=self._start_conversion,
        )
        self._btn_convert.pack(side="left", padx=(0, 10))

        self._btn_open = tk.Button(
            btn_frm, text="  Open Result  ",
            bg="#34A853", fg=self._BTN_FG,
            font=("Helvetica", 11),
            relief="flat", cursor="hand2", padx=14, pady=6,
            state="disabled",
            command=self._open_result,
        )
        self._btn_open.pack(side="left")

        # Log area
        tk.Label(body, text="Log", bg=self._BG,
                 font=("Helvetica", 10, "bold"), anchor="w").pack(fill="x")
        log_frm = tk.Frame(body, bg=self._BG)
        log_frm.pack(fill="both", expand=True)

        self._log_text = tk.Text(
            log_frm, height=8, state="disabled",
            bg="#1E1E1E", fg="#D4D4D4",
            font=("Courier", 9), relief="flat",
            wrap="word", padx=8, pady=6,
        )
        scroll = ttk.Scrollbar(log_frm, command=self._log_text.yview)
        self._log_text.configure(yscrollcommand=scroll.set)
        self._log_text.pack(side="left", fill="both", expand=True)
        scroll.pack(side="right", fill="y")

    def _center_window(self, w, h):
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x  = (sw - w) // 2
        y  = (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _log(self, msg: str):
        def _do():
            self._log_text.configure(state="normal")
            self._log_text.insert("end", msg + "\n")
            self._log_text.see("end")
            self._log_text.configure(state="disabled")
        self.after(0, _do)

    def _set_status(self, msg: str):
        self.after(0, lambda: self._status_msg.set(msg))

    def _set_progress(self, value: float):
        self.after(0, lambda: self._progress.configure(value=value))

    # ── Events ────────────────────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Select a PDF",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
        )
        if path:
            self._load_pdf(path)

    def _on_drop(self, event):
        # tkinterdnd2 wraps paths in {} on macOS/Windows for paths with spaces
        raw = event.data.strip()
        path = raw.strip("{}") if raw.startswith("{") else raw.split()[0]
        if path.lower().endswith(".pdf"):
            self._load_pdf(path)
        else:
            messagebox.showwarning("Not a PDF", "Please drop a PDF file.")

    def _load_pdf(self, path: str):
        self._pdf_path.set(path)
        self._output_path = None
        self._btn_open.configure(state="disabled")
        self._progress["value"] = 0
        self._set_status("Ready to convert.")
        self._log(f"Selected: {Path(path).name}")

    def _open_result(self):
        if self._output_path and Path(self._output_path).exists():
            _open_file(self._output_path)

    def _ask_combine(self, n: int) -> bool:
        """Ask the user (from the worker thread) whether to combine tables. Thread-safe."""
        result = [False]
        event = threading.Event()

        def _ask():
            result[0] = messagebox.askyesno(
                "Combine tables?",
                f"All {n} tables have identical headers.\n\nCombine them into one sheet?",
                icon="question",
            )
            event.set()

        self.after(0, _ask)
        event.wait()
        return result[0]

    # ── Conversion ───────────────────────────────────────────────────────────

    def _start_conversion(self):
        pdf = self._pdf_path.get()
        if not pdf:
            messagebox.showinfo("No file selected", "Please select a PDF file first.")
            return
        if not Path(pdf).exists():
            messagebox.showerror("File not found", f"Cannot find:\n{pdf}")
            return
        if self._converting:
            return

        self._converting = True
        self._btn_convert.configure(state="disabled", text="  Converting…  ")
        self._btn_open.configure(state="disabled")
        self._progress["value"] = 0
        self._set_status("Starting…")
        self._log("─" * 50)
        self._log(f"Converting: {Path(pdf).name}")

        thread = threading.Thread(target=self._run_conversion, args=(pdf,), daemon=True)
        thread.start()

    def _run_conversion(self, pdf_path: str):
        out_path = None
        try:
            def progress_cb(page, total):
                pct = int(page / total * 90)
                self._set_progress(pct)
                self._set_status(f"Processing page {page} of {total}…")

            tables = extract_tables_from_pdf(pdf_path, progress_cb, self._log)

            if not tables:
                self._log("⚠  No content found in this PDF.")
                self._set_status("Done — no tables or text found.")
                self._set_progress(100)
                return

            if _all_headers_match(tables):
                self._log(f"Detected {len(tables)} tables with identical headers.")
                if self._ask_combine(len(tables)):
                    tables = _combine_tables(tables)
                    self._log("Combined into 1 sheet.")
                else:
                    self._log(f"Keeping {len(tables)} separate sheets.")

            self._set_status("Writing Excel file…")
            self._log(f"Writing {len(tables)} sheet(s)…")

            out_path = _output_path(pdf_path)
            write_excel(tables, pdf_path, out_path)

            self._set_progress(100)
            self._output_path = out_path
            msg = f"Saved: {Path(out_path).name}  ({len(tables)} table(s))"
            self._set_status(msg)
            self._log(f"✓ {msg}")
            self._log(f"  Full path: {out_path}")
            self.after(0, lambda: self._btn_open.configure(state="normal"))

        except Exception as exc:
            err = traceback.format_exc()
            self._log(f"✗ Error: {exc}")
            self._log(err)
            self._set_status("Conversion failed — see log.")
            # Clean up partial file
            if out_path and Path(out_path).exists():
                try:
                    Path(out_path).unlink()
                except Exception:
                    pass

        finally:
            self._converting = False
            self.after(0, lambda: self._btn_convert.configure(
                state="normal", text="  Convert  "
            ))


# ── 6. Entry point ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    app = App()
    app.mainloop()
